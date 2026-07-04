"""Generate seed/facts_us.json from the BLS OEWS national wage file.

Real data only (honesty-first, DESIGN-EN Wave EN-3): annual US wages by SOC,
joined to our occupations via onet_code. No LLM, no made-up numbers.

Download the OEWS national file first (public, no key; needs a UA):

    curl -L --ssl-no-revoke -A "astrolab-research (you@example.com)" \\
      -o /tmp/oesnat.zip https://www.bls.gov/oes/special-requests/oesm24nat.zip

Then:  python -m scripts.fetch_us_facts --zip /tmp/oesnat.zip

Wages are ANNUAL USD (not monthly like RU). salary_low = 25th pct, salary_high
= 90th pct. BLS "#" (wage >= $239,200/yr) is capped at that value. Occupations
with no onet_code or no OEWS match are skipped and reported.
"""

from __future__ import annotations

import argparse
import io
import json
import zipfile
from pathlib import Path

import openpyxl

SEED_DIR = Path(__file__).resolve().parent.parent / "seed"
BATCHES = ["occupations_batch1.json", "occupations_batch2.json", "occupations_batch3.json"]
AS_OF = "2024"  # OEWS reference: May 2024
BLS_CAP = 239200  # BLS "#" = annual wage >= $239,200

# A few O*NET-SOC codes have no matching OEWS *detailed* wage row; map to the
# right OEWS series (real BLS numbers, just a code alignment — not a guess).
_SOC_REMAP = {"19-3031": "19-3033"}  # psychologist -> Clinical & Counseling Psychologists


def _num(v: object) -> int | None:
    """Parse an OEWS wage/emp cell: '#' -> cap, '*'/'' -> None, else int."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "*", "**"):
        return None
    if s == "#":
        return BLS_CAP
    try:
        return int(float(s))
    except ValueError:
        return None


def _human(emp: int) -> str:
    if emp >= 1_000_000:
        return f"{emp / 1_000_000:.2f}".rstrip("0").rstrip(".") + "M"
    if emp >= 1_000:
        return f"{round(emp / 1_000)}k"
    return str(emp)


def _load_onet_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for name in BATCHES:
        for occ in json.loads((SEED_DIR / name).read_text("utf-8")):
            if occ.get("onet_code"):
                mapping[occ["slug"]] = occ["onet_code"]
    return mapping


def _load_oews(zip_path: Path) -> dict[str, dict]:
    z = zipfile.ZipFile(zip_path)
    member = next(n for n in z.namelist() if n.endswith("_dl.xlsx"))
    wb = openpyxl.load_workbook(io.BytesIO(z.read(member)), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    idx = {str(h): i for i, h in enumerate(next(rows))}
    by_soc: dict[str, dict] = {}
    for r in rows:
        soc = str(r[idx["OCC_CODE"]]).strip()
        if soc in by_soc:  # first row per SOC (national cross-industry)
            continue
        by_soc[soc] = {
            "title": r[idx["OCC_TITLE"]],
            "emp": _num(r[idx["TOT_EMP"]]),
            "p25": _num(r[idx["A_PCT25"]]),
            "median": _num(r[idx["A_MEDIAN"]]),
            "p90": _num(r[idx["A_PCT90"]]),
        }
    return by_soc


def run(zip_path: Path) -> None:
    onet = _load_onet_map()
    oews = _load_oews(zip_path)

    facts: list[dict] = []
    skipped: list[str] = []
    for slug, onet_code in sorted(onet.items()):
        soc = onet_code.split(".")[0]  # 15-1252.00 -> 15-1252
        soc = _SOC_REMAP.get(soc, soc)
        row = oews.get(soc)
        if row is None or row["p25"] is None or row["p90"] is None:
            skipped.append(f"{slug} (onet {onet_code} / soc {soc})")
            continue
        emp = row["emp"]
        demand = f"~{_human(emp)} jobs in the US (BLS OEWS, May 2024)" if emp else None
        facts.append(
            {
                "slug": slug,
                "soc": soc,
                "bls_title": row["title"],
                "salary_low": row["p25"],
                "salary_high": row["p90"],
                "median": row["median"],
                "currency": "USD",
                "period": "year",
                "demand_note": demand,
                "as_of": AS_OF,
            }
        )

    out = {
        "country": "US",
        "source": "BLS OEWS national, May 2024 (bls.gov/oes)",
        "as_of": AS_OF,
        "facts": facts,
    }
    (SEED_DIR / "facts_us.json").write_text(
        json.dumps(out, indent=2, ensure_ascii=False) + "\n", "utf-8"
    )
    print(f"[fetch_us_facts] wrote {len(facts)} facts to seed/facts_us.json")
    if skipped:
        print(f"[fetch_us_facts] skipped {len(skipped)}: " + "; ".join(skipped))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", required=True, help="path to OEWS national zip (oesmNNnat.zip)")
    ap.add_argument("--out", default=None)  # reserved
    args = ap.parse_args()
    run(Path(args.zip))


if __name__ == "__main__":
    main()
